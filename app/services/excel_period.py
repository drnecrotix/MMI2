from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


MONTH_NAMES = {
    "януари": 1, "ян": 1, "january": 1, "jan": 1,
    "февруари": 2, "фев": 2, "february": 2, "feb": 2,
    "март": 3, "mar": 3, "march": 3,
    "април": 4, "апр": 4, "april": 4, "apr": 4,
    "май": 5, "may": 5,
    "юни": 6, "jun": 6, "june": 6,
    "юли": 7, "jul": 7, "july": 7,
    "август": 8, "авг": 8, "august": 8, "aug": 8,
    "септември": 9, "сеп": 9, "september": 9, "sep": 9,
    "октомври": 10, "окт": 10, "october": 10, "oct": 10,
    "ноември": 11, "ное": 11, "november": 11, "nov": 11,
    "декември": 12, "дек": 12, "december": 12, "dec": 12,
}


@dataclass(frozen=True)
class PeriodDetection:
    year: int
    month: int
    confidence: str
    evidence: str


def _norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def _text_candidates(text: str) -> list[tuple[int, int]]:
    out: list[tuple[int, int]] = []
    normalized = _norm(text)
    if not normalized:
        return out

    for match in re.finditer(r"(?<!\d)(0?[1-9]|1[0-2])[./-](20\d{2})(?!\d)", normalized):
        out.append((int(match.group(2)), int(match.group(1))))
    for match in re.finditer(r"(?<!\d)(20\d{2})[./-](0?[1-9]|1[0-2])(?!\d)", normalized):
        out.append((int(match.group(1)), int(match.group(2))))

    years = [int(y) for y in re.findall(r"(?<!\d)(20\d{2})(?!\d)", normalized)]
    if years:
        for name, month in MONTH_NAMES.items():
            if re.search(rf"(?<!\w){re.escape(name)}(?!\w)", normalized):
                out.extend((year, month) for year in years)
    return out


def detect_schedule_period(content: bytes, filename: str = "") -> PeriodDetection | None:
    if filename and Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("Поддържат се само .xlsx файлове.")

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    scores: Counter[tuple[int, int]] = Counter()
    evidence: dict[tuple[int, int], list[str]] = {}

    def add(period: tuple[int, int], weight: int, why: str) -> None:
        year, month = period
        if 2020 <= year <= 2100 and 1 <= month <= 12:
            scores[period] += weight
            evidence.setdefault(period, []).append(why)

    for period in _text_candidates(filename):
        add(period, 2, "име на файла")

    for ws in workbook.worksheets:
        for period in _text_candidates(ws.title):
            add(period, 2, f"име на листа {ws.title}")

        max_row = min(ws.max_row or 1, 50)
        max_col = min(ws.max_column or 1, 60)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                value = cell.value
                if isinstance(value, datetime):
                    add((value.year, value.month), 5, f"дата в {ws.title}!{cell.coordinate}")
                elif isinstance(value, date):
                    add((value.year, value.month), 5, f"дата в {ws.title}!{cell.coordinate}")
                elif isinstance(value, str):
                    for period in _text_candidates(value):
                        add(period, 3, f"текст в {ws.title}!{cell.coordinate}")

    if not scores:
        return None

    ranked = scores.most_common(2)
    best_period, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0
    if best_score < 2 or (second_score and best_score == second_score):
        return None

    confidence = "high" if best_score >= 5 and best_score >= second_score + 2 else "medium"
    unique_evidence = list(dict.fromkeys(evidence.get(best_period, [])))[:3]
    return PeriodDetection(
        year=best_period[0],
        month=best_period[1],
        confidence=confidence,
        evidence=", ".join(unique_evidence) or "Excel metadata",
    )
