from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Employee, ShiftEntry


WORK_NUMBER_HEADERS = {
    "раб. №",
    "раб №",
    "работен номер",
    "работен №",
    "работен no",
    "табелен номер",
    "таб. номер",
    "employee number",
    "work number",
    "id",
}
NAME_HEADERS = {
    "име, фамилия",
    "име фамилия",
    "име",
    "име на служителя",
    "служител",
    "трите имена",
    "name",
    "employee",
}

# Employee-cell codes observed in the real MMI2 workbook.
# Numeric work codes are classified as day/night from the schedule-pattern row,
# because the same code can occur on different shift dates.
DIRECT_SHIFT_ALIASES = {
    "О": "leave",
    "ОТП": "leave",
    "ОТПУСК": "leave",
    "LEAVE": "leave",
    "VACATION": "leave",
    "Б": "sick_leave",
    "БОЛ": "sick_leave",
    "БОЛНИЧЕН": "sick_leave",
    "SICK": "sick_leave",
    "К": "compensation",
    "КОМП": "compensation",
    "КОМПЕНСАЦИЯ": "compensation",
    "COMP": "compensation",
    "П": "rest",
    "ПОЧ": "rest",
    "ПОЧИВКА": "rest",
    "REST": "rest",
    "OFF": "rest",
}
WORK_CODES = {"1", "2", "8", "I", "І"}
NIGHT_PATTERN_MARKERS = {"Н", "N", "NIGHT", "НОЩ", "НОЩНА"}
DAY_PATTERN_MARKERS = {"Д", "D", "DAY", "ДЕН", "ДНЕВНА", "А", "Б", "В", "Г"}


@dataclass
class ImportResult:
    employees: int = 0
    shifts: int = 0
    skipped_rows: int = 0
    schedule_blocks: int = 0


@dataclass
class ScheduleBlock:
    header_row: int
    days_row: int
    pattern_row: int
    first_employee_row: int
    last_employee_row: int
    work_col: int
    name_col: int
    day_columns: dict[int, int]
    day_patterns: dict[int, str]


def _norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _raw(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _code(value: object) -> str:
    return _raw(value).upper().replace(".", "").strip()


def _day_from_header(value: object, year: int, month: int) -> int | None:
    if isinstance(value, datetime):
        return value.day if value.year == year and value.month == month else None
    if isinstance(value, date):
        return value.day if value.year == year and value.month == month else None
    if isinstance(value, int) and 1 <= value <= 31:
        return value
    if isinstance(value, float) and value.is_integer() and 1 <= int(value) <= 31:
        return int(value)
    text = _norm(value)
    if text.isdigit() and 1 <= int(text) <= 31:
        return int(text)
    return None


def _find_header_columns(ws, row_idx: int) -> tuple[int, int] | None:
    normalized = {
        _norm(ws.cell(row_idx, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row_idx, col).value is not None
    }
    work_col = next((normalized[h] for h in WORK_NUMBER_HEADERS if h in normalized), None)
    name_col = next((normalized[h] for h in NAME_HEADERS if h in normalized), None)
    if work_col and name_col:
        return work_col, name_col
    return None


def _find_days_row(ws, header_row: int, year: int, month: int) -> tuple[int, dict[int, int]] | None:
    # MMI2 uses a merged "Дати на месеца" header and a separate 1..31 row.
    # Search a few rows below the employee-column header to support small layout changes.
    for row_idx in range(header_row, min(header_row + 5, ws.max_row) + 1):
        day_columns: dict[int, int] = {}
        for col in range(1, ws.max_column + 1):
            day = _day_from_header(ws.cell(row_idx, col).value, year, month)
            if day is not None:
                day_columns[col] = day
        if len(day_columns) >= 20:
            return row_idx, day_columns
    return None


def _is_employee_row(ws, row: int, work_col: int, name_col: int) -> bool:
    work_number = _raw(ws.cell(row, work_col).value)
    full_name = _raw(ws.cell(row, name_col).value)
    if not work_number or not full_name:
        return False
    # Avoid accidentally treating headings/formulas as employees.
    return any(ch.isdigit() for ch in work_number) and len(full_name) >= 3


def _find_employee_bounds(ws, start_row: int, work_col: int, name_col: int) -> tuple[int, int] | None:
    first: int | None = None
    last: int | None = None
    empty_run = 0

    for row in range(start_row, ws.max_row + 1):
        if _find_header_columns(ws, row) and first is not None:
            break

        if _is_employee_row(ws, row, work_col, name_col):
            if first is None:
                first = row
            last = row
            empty_run = 0
            continue

        if first is not None:
            empty_run += 1
            # Real blocks are followed by summaries/signatures. A small gap inside
            # the employee list is tolerated, but a longer run closes the block.
            if empty_run >= 4:
                break

    if first is None or last is None:
        return None
    return first, last


def _discover_blocks(ws, year: int, month: int) -> list[ScheduleBlock]:
    blocks: list[ScheduleBlock] = []
    row = 1

    while row <= ws.max_row:
        columns = _find_header_columns(ws, row)
        if not columns:
            row += 1
            continue

        work_col, name_col = columns
        days_info = _find_days_row(ws, row, year, month)
        if not days_info:
            row += 1
            continue

        days_row, day_columns = days_info
        pattern_row = days_row + 2
        # In some exports weekday is directly below dates and pattern is the next row.
        # If the first candidate has no markers, try the immediately following row.
        pattern_values = [_code(ws.cell(pattern_row, col).value) for col in day_columns]
        if not any(v in NIGHT_PATTERN_MARKERS or v in DAY_PATTERN_MARKERS for v in pattern_values):
            pattern_row = days_row + 1

        bounds = _find_employee_bounds(ws, pattern_row + 1, work_col, name_col)
        if not bounds:
            row += 1
            continue
        first_employee_row, last_employee_row = bounds

        day_patterns = {
            col: _code(ws.cell(pattern_row, col).value)
            for col in day_columns
        }
        blocks.append(
            ScheduleBlock(
                header_row=row,
                days_row=days_row,
                pattern_row=pattern_row,
                first_employee_row=first_employee_row,
                last_employee_row=last_employee_row,
                work_col=work_col,
                name_col=name_col,
                day_columns=day_columns,
                day_patterns=day_patterns,
            )
        )
        row = last_employee_row + 1

    return blocks


def _shift_type(value: object, pattern_marker: str) -> tuple[str, str]:
    raw_code = _raw(value)
    code = _code(value)

    if code in DIRECT_SHIFT_ALIASES:
        return DIRECT_SHIFT_ALIASES[code], raw_code

    if code in WORK_CODES:
        if pattern_marker in NIGHT_PATTERN_MARKERS:
            return "night", raw_code
        # The MMI2 pattern row uses Д for day in most blocks and occasionally the
        # team letter (for example Б) in the same position. Both mean a day shift.
        if pattern_marker in DAY_PATTERN_MARKERS or pattern_marker:
            return "day", raw_code
        return "work", raw_code

    if not code:
        # Empty schedule cells are meaningful in this workbook: they are regular
        # days off between shifts and must be returned by the mobile/web calendar.
        return "rest", ""

    return "unknown", raw_code


def import_schedule_xlsx(db: Session, content: bytes, filename: str, year: int, month: int) -> ImportResult:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("Поддържат се само .xlsx файлове.")

    workbook = load_workbook(BytesIO(content), data_only=True)
    result = ImportResult()
    seen_employees: set[int] = set()

    for ws in workbook.worksheets:
        blocks = _discover_blocks(ws, year, month)
        result.schedule_blocks += len(blocks)

        for block in blocks:
            for row in range(block.first_employee_row, block.last_employee_row + 1):
                if not _is_employee_row(ws, row, block.work_col, block.name_col):
                    result.skipped_rows += 1
                    continue

                work_number = _raw(ws.cell(row, block.work_col).value)
                full_name = _raw(ws.cell(row, block.name_col).value)

                employee = db.scalar(select(Employee).where(Employee.work_number == work_number))
                if employee is None:
                    employee = Employee(work_number=work_number, full_name=full_name)
                    db.add(employee)
                    db.flush()
                elif employee.full_name != full_name:
                    employee.full_name = full_name

                seen_employees.add(employee.id)

                for col, day in block.day_columns.items():
                    try:
                        work_date = date(year, month, day)
                    except ValueError:
                        continue

                    shift_type, raw_code = _shift_type(
                        ws.cell(row, col).value,
                        block.day_patterns.get(col, ""),
                    )
                    existing = db.scalar(
                        select(ShiftEntry).where(
                            ShiftEntry.employee_id == employee.id,
                            ShiftEntry.work_date == work_date,
                        )
                    )
                    if existing:
                        existing.shift_type = shift_type
                        existing.raw_code = raw_code
                        existing.source_file = filename
                    else:
                        db.add(
                            ShiftEntry(
                                employee_id=employee.id,
                                work_date=work_date,
                                shift_type=shift_type,
                                raw_code=raw_code,
                                source_file=filename,
                            )
                        )
                    result.shifts += 1

    db.commit()
    result.employees = len(seen_employees)
    if result.employees == 0:
        raise ValueError("Не беше открит използваем MMI2 график с данни за служители.")
    return result
