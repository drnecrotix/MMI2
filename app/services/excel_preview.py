from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.services.excel_import import _code, _discover_blocks, _is_employee_row, _raw, _shift_type


@dataclass
class EmployeePreview:
    work_number: str
    full_name: str
    team: str
    totals: Counter = field(default_factory=Counter)


@dataclass
class PreviewResult:
    employees: list[dict]
    schedule_blocks: int
    duplicate_employee_rows: int
    conflicting_days: int
    unknown_codes: dict[str, int]
    totals: dict[str, int]


def preview_schedule_xlsx(content: bytes, filename: str, year: int, month: int) -> PreviewResult:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("Поддържат се само .xlsx файлове.")

    workbook = load_workbook(BytesIO(content), data_only=True)
    employees: dict[str, EmployeePreview] = {}
    seen_rows: set[str] = set()
    seen_days: dict[tuple[str, date], tuple[str, str]] = {}
    unknown_codes: Counter[str] = Counter()
    global_totals: Counter[str] = Counter()
    schedule_blocks = 0
    duplicate_employee_rows = 0
    conflicting_days = 0

    for ws in workbook.worksheets:
        blocks = _discover_blocks(ws, year, month)
        schedule_blocks += len(blocks)

        for block in blocks:
            for row in range(block.first_employee_row, block.last_employee_row + 1):
                if not _is_employee_row(ws, row, block.work_col, block.name_col, block.team_col):
                    continue

                work_number = _raw(ws.cell(row, block.work_col).value)
                full_name = _raw(ws.cell(row, block.name_col).value)
                team = _code(ws.cell(row, block.team_col).value)

                if work_number in seen_rows:
                    duplicate_employee_rows += 1
                seen_rows.add(work_number)

                employee = employees.get(work_number)
                if employee is None:
                    employee = EmployeePreview(work_number=work_number, full_name=full_name, team=team)
                    employees[work_number] = employee
                else:
                    employee.full_name = full_name
                    employee.team = team

                for col, day in block.day_columns.items():
                    try:
                        work_date = date(year, month, day)
                    except ValueError:
                        continue

                    shift_type, raw_code = _shift_type(ws.cell(row, col).value)
                    key = (work_number, work_date)
                    previous = seen_days.get(key)
                    current = (shift_type, raw_code)
                    if previous is not None and previous != current:
                        conflicting_days += 1
                    seen_days[key] = current

    # Build totals from the final value for every employee/date, mirroring import overwrite behavior.
    for (work_number, _), (shift_type, raw_code) in seen_days.items():
        employee = employees[work_number]
        employee.totals[shift_type] += 1
        global_totals[shift_type] += 1
        if shift_type == "unknown" and raw_code:
            unknown_codes[raw_code] += 1

    if not employees:
        raise ValueError("Не беше открит използваем MMI2 график с данни за служители.")

    employee_rows = [
        {
            "work_number": employee.work_number,
            "full_name": employee.full_name,
            "team": employee.team,
            "day": employee.totals.get("day", 0),
            "night": employee.totals.get("night", 0),
            "leave": employee.totals.get("leave", 0),
            "sick_leave": employee.totals.get("sick_leave", 0),
            "rest": employee.totals.get("rest", 0),
            "unknown": employee.totals.get("unknown", 0),
        }
        for employee in sorted(employees.values(), key=lambda item: (item.team, item.full_name, item.work_number))
    ]

    return PreviewResult(
        employees=employee_rows,
        schedule_blocks=schedule_blocks,
        duplicate_employee_rows=duplicate_employee_rows,
        conflicting_days=conflicting_days,
        unknown_codes=dict(sorted(unknown_codes.items())),
        totals=dict(global_totals),
    )
