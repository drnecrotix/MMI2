from io import BytesIO
import unittest

from openpyxl import Workbook

from app.services.excel_preview import preview_schedule_xlsx


class ExcelPreviewTests(unittest.TestCase):
    def _workbook(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws["C1"] = "Раб. №"
        ws["D1"] = "Име, фамилия"
        for index, day in enumerate(range(1, 32), start=7):
            ws.cell(2, index).value = day
        ws["F3"] = "А"
        ws["C3"] = 1234
        ws["D3"] = "Тест Служител"
        ws["G3"] = 1
        ws["H3"] = 2
        ws["I3"] = "О"
        ws["J3"] = "Б"
        ws["K3"] = "X"
        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def test_preview_detects_employee_and_codes_without_database(self):
        result = preview_schedule_xlsx(self._workbook(), "schedule.xlsx", 2026, 7)
        self.assertEqual(len(result.employees), 1)
        employee = result.employees[0]
        self.assertEqual(employee["work_number"], "1234")
        self.assertEqual(employee["team"], "А")
        self.assertEqual(employee["day"], 1)
        self.assertEqual(employee["night"], 1)
        self.assertEqual(employee["leave"], 1)
        self.assertEqual(employee["sick_leave"], 1)
        self.assertEqual(employee["unknown"], 1)
        self.assertEqual(result.unknown_codes, {"X": 1})


if __name__ == "__main__":
    unittest.main()
