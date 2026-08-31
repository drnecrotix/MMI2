from io import BytesIO
import unittest
from unittest.mock import patch

from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from app.db import Base
from app.models import Employee, ShiftEntry
from app.services.excel_import import import_schedule_xlsx


class ImportTransactionTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite+pysqlite:///:memory:")
        Base.metadata.create_all(self.engine)

    @staticmethod
    def _valid_workbook() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws["C1"] = "Раб. №"
        ws["D1"] = "Име, фамилия"
        for day, col in enumerate(range(7, 38), start=1):
            ws.cell(2, col).value = day
        ws["C3"] = 1234
        ws["D3"] = "Тест Служител"
        ws["F3"] = "А"
        ws["G3"] = 1
        ws["H3"] = 2

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    @staticmethod
    def _invalid_workbook() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "няма график"
        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def test_importer_never_commits_and_can_be_rolled_back(self):
        with Session(self.engine) as db:
            with patch.object(db, "commit", wraps=db.commit) as commit:
                result = import_schedule_xlsx(db, self._valid_workbook(), "schedule.xlsx", 2026, 7)
                commit.assert_not_called()

            self.assertEqual(result.employees, 1)
            self.assertEqual(result.shifts, 31)
            self.assertEqual(db.scalar(select(func.count()).select_from(Employee)), 1)
            self.assertEqual(db.scalar(select(func.count()).select_from(ShiftEntry)), 31)

            db.rollback()

            self.assertEqual(db.scalar(select(func.count()).select_from(Employee)), 0)
            self.assertEqual(db.scalar(select(func.count()).select_from(ShiftEntry)), 0)

    def test_invalid_workbook_does_not_commit_before_error(self):
        with Session(self.engine) as db:
            with patch.object(db, "commit", wraps=db.commit) as commit:
                with self.assertRaises(ValueError):
                    import_schedule_xlsx(db, self._invalid_workbook(), "invalid.xlsx", 2026, 7)
                commit.assert_not_called()

            db.rollback()
            self.assertEqual(db.scalar(select(func.count()).select_from(Employee)), 0)
            self.assertEqual(db.scalar(select(func.count()).select_from(ShiftEntry)), 0)


if __name__ == "__main__":
    unittest.main()
