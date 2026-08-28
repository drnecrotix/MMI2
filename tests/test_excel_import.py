from io import BytesIO
import unittest

from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.db import Base
from app.models import Employee, ShiftEntry
from app.services.excel_import import import_schedule_xlsx


class ExcelImportTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite+pysqlite:///:memory:")
        Base.metadata.create_all(self.engine)

    def _workbook_bytes(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "ГРАФИК"

        ws["B1"] = "№ ред"
        ws["C1"] = "Раб. №"
        ws["D1"] = "Име, фамилия"
        ws["E1"] = "длъжност"
        ws["G2"] = 1
        ws["H2"] = 2
        ws["I2"] = 3
        ws["J2"] = 4
        ws["K2"] = 5
        ws["L2"] = 6
        ws["M2"] = 7
        ws["N2"] = 8
        ws["O2"] = 9
        ws["P2"] = 10
        ws["Q2"] = 11
        ws["R2"] = 12
        ws["S2"] = 13
        ws["T2"] = 14
        ws["U2"] = 15
        ws["V2"] = 16
        ws["W2"] = 17
        ws["X2"] = 18
        ws["Y2"] = 19
        ws["Z2"] = 20
        ws["AA2"] = 21
        ws["AB2"] = 22
        ws["AC2"] = 23
        ws["AD2"] = 24
        ws["AE2"] = 25
        ws["AF2"] = 26
        ws["AG2"] = 27
        ws["AH2"] = 28
        ws["AI2"] = 29
        ws["AJ2"] = 30
        ws["AK2"] = 31

        # weekday row
        for col in range(7, 38):
            ws.cell(3, col).value = "Ср"

        # shift-pattern row
        ws["G4"] = "Д"
        ws["H4"] = "Н"
        ws["I4"] = "Д"
        ws["J4"] = "Н"

        ws["B5"] = 1
        ws["C5"] = 1234
        ws["D5"] = "Тест Служител"
        ws["E5"] = "маш. лок."
        ws["G5"] = 1
        ws["H5"] = 2
        ws["I5"] = "О"
        ws["J5"] = "Б"
        ws["K5"] = "К"
        ws["L5"] = "П"
        # Gaps after that are intentional regular rest days.

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def test_real_mmi2_codes_and_empty_rest_days(self):
        with Session(self.engine) as db:
            result = import_schedule_xlsx(db, self._workbook_bytes(), "test.xlsx", 2026, 7)
            self.assertEqual(result.schedule_blocks, 1)
            self.assertEqual(result.employees, 1)
            self.assertEqual(result.shifts, 31)

            employee = db.scalar(select(Employee).where(Employee.work_number == "1234"))
            entries = db.scalars(
                select(ShiftEntry)
                .where(ShiftEntry.employee_id == employee.id)
                .order_by(ShiftEntry.work_date)
            ).all()
            by_day = {entry.work_date.day: entry for entry in entries}

            self.assertEqual(by_day[1].shift_type, "day")
            self.assertEqual(by_day[2].shift_type, "night")
            self.assertEqual(by_day[3].shift_type, "leave")
            self.assertEqual(by_day[4].shift_type, "sick_leave")
            self.assertEqual(by_day[5].shift_type, "compensation")
            self.assertEqual(by_day[6].shift_type, "rest")
            self.assertEqual(by_day[7].shift_type, "rest")
            self.assertEqual(by_day[7].raw_code, "")


if __name__ == "__main__":
    unittest.main()
